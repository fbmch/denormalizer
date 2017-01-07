
import openpyxl


class Xlsx:


    fonts = None
    sheetnames = None

    def dump(filename, table):
    
        wb = openpyxl.Workbook(write_only=True)

        ws = wb.create_sheet()

        for r in table:
            ws.append(r)

        wb._fonts = Xlsx.fonts 
        wb.save(filename)


    def load(filename):
        
        wb = openpyxl.load_workbook(filename=filename, read_only=True)
        Xlsx.fonts = wb._fonts
    
        return [ (ws.title, Xlsx._rows(ws)) for ws in wb._sheets ]
   

    def _rows(sheet):

        return [ [cell.value for cell in row] for row in sheet.rows]
